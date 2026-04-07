from __future__ import annotations

import base64
import io
import logging
from datetime import date, timedelta
from pathlib import Path
from statistics import median
from typing import Optional

import pandas as pd
import streamlit as st

from bam_curve_fetcher import BamCurveFetcher
from vba_equivalent_rates import calcul_taux

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
try:
    CREDENTIALS = {
        st.secrets["auth"]["username"]: st.secrets["auth"]["password"]
    }
except Exception:
    st.error("Secrets non configurés. Contactez l'administrateur.")
    st.stop()
APP_TITLE   = "Spread Manager — BSF & CD | Al Barid Bank"
CACHE_DIR   = Path(__file__).parent / "cache_bam_curves"
ASSETS_DIR  = Path(__file__).parent / "assets"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# BANK ISSUERS TO EXCLUDE (OBLIG_ORDN)
# ─────────────────────────────────────────────────────────────────────────────
BANQUES_EXCLUES = {
    'ATW E', 'BCP E', 'CIH E', 'BOA', 'CAM E', 'CDM', 'CDG E',
    'CFG BANK', 'CFC', 'CFM', 'SAHAM FINANCES', 'BMCI',
    'BMCI LEASI', 'AL BARID BANK E', 'CREDIPER FT',
}
# Alias pour compatibilité
BANK_ISSUERS = BANQUES_EXCLUES

# ─────────────────────────────────────────────────────────────────────────────
# SECTOR MAP (OBLIG_ORDN) — matching on PREFERREDNAMEISSUER exact
# ─────────────────────────────────────────────────────────────────────────────
SECTEUR_MAP = {
    # ENERGIE
    'ONEE': 'ENERGIE', 'TAQA MOROCCO': 'ENERGIE', 'MASEN': 'ENERGIE',
    'FEC': 'ENERGIE', 'LYDEC': 'ENERGIE', 'SAMIR': 'ENERGIE',
    'SEDM': 'ENERGIE', 'TMPA': 'ENERGIE', 'AFRIQUIA GAZ': 'ENERGIE',
    'AFIQUIA LUB': 'ENERGIE', 'MAGHREB OXYGENE': 'ENERGIE',
    'PHOSPH BOUCRAA': 'ENERGIE',
    # BTP
    'TGCC': 'BTP', 'JET CONTRACTORS': 'BTP', 'ADI': 'BTP',
    'SETTAVEX': 'BTP', 'CMGP GROUP': 'BTP', 'ASSIFILL BUILD': 'BTP',
    'BUILDING LOGIST': 'BTP', 'NADOR WEST MED': 'BTP',
    # CIMENTERIE
    'HOLCIM MAROC': 'CIMENTERIE', 'CIMAT': 'CIMENTERIE',
    # IMMOBILIER
    'CGI': 'IMMOBILIER', 'ALLIANCES DARNA': 'IMMOBILIER',
    'DOUJA PROM ADD': 'IMMOBILIER', 'AL OMRANE': 'IMMOBILIER',
    'IMMOLOG': 'IMMOBILIER', 'ARADEI CAPITAL': 'IMMOBILIER',
    'BEST REAL ESTAT': 'IMMOBILIER', 'HAYAN IMMO SA': 'IMMOBILIER',
    'PALME DEV': 'IMMOBILIER', 'CMR LSTONE': 'IMMOBILIER',
    'Immorent SC': 'IMMOBILIER', 'RDS': 'IMMOBILIER',
    # TRANSPORT
    'ONCF': 'TRANSPORT', 'ONDA': 'TRANSPORT', 'CTM SA': 'TRANSPORT',
    'COMANAV': 'TRANSPORT', 'TANGER MED  SA': 'TRANSPORT',
    'TMSA': 'TRANSPORT', 'ANP': 'TRANSPORT', 'ADM': 'TRANSPORT',
    'MEDIACO MAROC': 'TRANSPORT', 'CMT': 'TRANSPORT',
    # TELECOM
    'MEDI TELCOM SA': 'TELECOM', 'IAM': 'TELECOM',
    # CHIMIE / MINES
    'OCP SA': 'CHIMIE / MINES', 'OCP NUTRICROPS': 'CHIMIE / MINES',
    'MANAGEM': 'CHIMIE / MINES',
    # AGRO-ALIMENTAIRE
    'OULMES': 'AGRO-ALIMENTAIRE', 'UNIMER': 'AGRO-ALIMENTAIRE',
    'AFRICA FEED FOOD': 'AGRO-ALIMENTAIRE',
    'SOMACOVAM': 'AGRO-ALIMENTAIRE', 'MUTANDIS SCA': 'AGRO-ALIMENTAIRE',
    'ZALAGH HOLDING': 'AGRO-ALIMENTAIRE',
    # SIDERURGIE
    'MAGHREB STEEL': 'SIDERURGIE',
    # SANTE
    'AKDITAL': 'SANTE',
    # TOURISME
    'RISMA SA': 'TOURISME', 'MADAEF': 'TOURISME',
    # GRANDE DISTRIBUTION
    'LABEL VIE': 'GRANDE DISTRIBUTION',
    'MARJANE HOLDING': 'GRANDE DISTRIBUTION',
    'RETAIL HOLDING': 'GRANDE DISTRIBUTION',
    # HOLDING
    'AL MADA': 'HOLDING', 'HOLMARCOM FIN CO': 'HOLDING',
    'FINANCECOM': 'HOLDING', 'FINANCIER SEC': 'HOLDING',
    'O CAPITAL GROUP': 'HOLDING', 'GROUP INVEST SA': 'HOLDING',
    'GARAN': 'HOLDING', 'Hold GENE EDU': 'HOLDING',
    # LEASING
    'MAGHREBAIL': 'LEASING', 'MA LEASING': 'LEASING',
    # CREDIT CONSO
    'WAFASALAF': 'CREDIT CONSO', 'TASLIF': 'CREDIT CONSO',
    # ASSURANCE
    'SAHAM': 'ASSURANCE',
    # MICROFINANCE
    'JAIDA': 'MICROFINANCE', 'SFI': 'MICROFINANCE',
    # PARA-PUBLIC
    'SCIF': 'PARA-PUBLIC', 'ORMVAD': 'PARA-PUBLIC',
    'COMMUNE AGADIR': 'PARA-PUBLIC', 'BEST FINANCIERE': 'PARA-PUBLIC',
    'AGRI CAPITAL': 'PARA-PUBLIC',
    # AUTOMOBILE
    'RCI': 'AUTOMOBILE', 'UNIVERS MOTORS': 'AUTOMOBILE',
    'AUTO NEJMA': 'AUTOMOBILE', 'OD MAROC': 'AUTOMOBILE',
    # EDUCATION
    'Univ Mohammed VI': 'EDUCATION',
    # TEXTILE
    'DISTRA-S.A': 'TEXTILE', 'MCM': 'TEXTILE',
    # INDUSTRIE
    'CMB PLA MAROC': 'INDUSTRIE', 'TC3PC': 'INDUSTRIE',
    'FINAN HATT': 'INDUSTRIE',
}

# ─────────────────────────────────────────────────────────────────────────────
# CSS  —  palette Al Barid Bank (thème sombre, orange brûlé)
# ─────────────────────────────────────────────────────────────────────────────
STYLE = """
<style>
[data-testid="stAppViewContainer"] {
    background: #100804 !important;
    color: #EDE0D4 !important;
}
[data-testid="stAppViewContainer"] > .main {
    background: #100804 !important;
    color: #EDE0D4 !important;
}
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1E0A03 0%, #130602 100%) !important;
}
[data-testid="stSidebar"] * { color: #EDE0D4 !important; }

/* widget labels */
label, p,
[data-testid="stWidgetLabel"] p,
[data-testid="stDateInput"] label,
[data-testid="stDateInput"] p,
[data-testid="stNumberInput"] label,
[data-testid="stNumberInput"] p,
[data-testid="stSelectbox"] label,
[data-testid="stSelectbox"] p,
[data-testid="stCheckbox"] p,
.stDateInput label, .stNumberInput label,
.stSelectbox label, .stCheckbox label {
    color: #F0EAE2 !important;
}

/* inputs */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stDateInput"] input {
    background: rgba(255,255,255,0.06) !important;
    color: #F0EAE2 !important;
    border: 1px solid rgba(200,80,30,0.35) !important;
    border-radius: 8px !important;
}
[data-baseweb="select"] > div {
    background: rgba(255,255,255,0.06) !important;
    color: #EDE0D4 !important;
    border: 1px solid rgba(200,80,30,0.35) !important;
}
[data-baseweb="popover"],
[data-baseweb="menu"],
[data-baseweb="option"] { background: #1E0A03 !important; color: #EDE0D4 !important; }
[data-baseweb="option"]:hover { background: rgba(200,80,30,0.25) !important; }
[data-testid="stCheckbox"] label { color: #EDE0D4 !important; }

/* file uploader */
[data-testid="stFileUploaderDropzone"] {
    background: rgba(255,255,255,0.05) !important;
    border: 1px dashed rgba(200,80,30,0.5) !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploaderDropzone"] * { color: #EDE0D4 !important; }

/* dataframe */
[data-testid="stDataFrame"] {
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid rgba(200,80,30,0.2) !important;
    border-radius: 10px !important;
}

/* progress */
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #C8501E, #FF9060) !important;
}

/* alerts */
.stAlert   { background: rgba(200,80,30,0.10) !important; border: 1px solid rgba(200,80,30,0.35) !important; }
.stInfo    { background: rgba(80,120,200,0.10) !important; border: 1px solid rgba(80,120,200,0.3) !important; }
.stSuccess { background: rgba(60,160,60,0.10)  !important; border: 1px solid rgba(60,160,60,0.3) !important; }
[data-testid="stAlertContainer"] * { color: #EDE0D4 !important; }

/* nav buttons */
.nav-btn button {
    background: linear-gradient(90deg, #C8501E 0%, #7A2E08 100%) !important;
    color: #FFFFFF !important; border: none !important;
    border-radius: 10px !important; font-weight: 700 !important;
    font-size: 14px !important; padding: 10px !important;
    margin-bottom: 6px !important; width: 100% !important;
}
.nav-btn button:hover { opacity: 0.88 !important; }
.nav-btn-active button {
    background: linear-gradient(90deg, #E07030 0%, #C8501E 100%) !important;
    color: #FFFFFF !important; border: none !important;
    border-radius: 10px !important; font-weight: 700 !important;
    font-size: 14px !important; padding: 10px !important;
    margin-bottom: 6px !important; width: 100% !important;
    box-shadow: 0 0 0 2px #F5C518 !important;
}

/* generic buttons */
.stButton > button {
    background: linear-gradient(90deg, #C8501E 0%, #7A2E08 100%) !important;
    color: #FFFFFF !important; border: none !important;
    border-radius: 10px !important; font-weight: 700 !important;
    padding: 10px !important; width: 100% !important;
}
.stButton > button:hover { opacity: 0.88 !important; }
.stDownloadButton > button {
    background: linear-gradient(90deg, #1A6B2E 0%, #0D4019 100%) !important;
    color: white !important; border: none !important;
    border-radius: 10px !important; font-weight: 700 !important;
    padding: 10px !important; width: 100% !important;
}
.stDownloadButton > button:hover { opacity: 0.88 !important; }

/* KPI cards */
.kpi-card {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(200,80,30,0.3);
    border-top: 3px solid #C8501E;
    border-radius: 10px; padding: 14px 16px;
    text-align: center; margin-bottom: 10px;
}
.kpi-title { font-size: 11px; color: #FFFFFF; font-weight: 600;
             letter-spacing: 0.8px; text-transform: uppercase; }
.kpi-value { font-size: 1.45rem; font-weight: 800; color: #FF9060; }
.kpi-sub   { font-size: 0.75rem; color: #FFFFFF; margin-top: 2px; }

/* section headers */
.section-hdr {
    background: linear-gradient(90deg, #C8501E 0%, #7A2E08 100%);
    color: #FFFFFF !important; padding: 9px 18px;
    border-radius: 8px; font-weight: 700; font-size: 0.95rem;
    margin: 20px 0 12px 0;
}

/* instruction cards */
.instr-card {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(200,80,30,0.25);
    border-left: 4px solid #C8501E;
    border-radius: 10px; padding: 16px 20px; margin-bottom: 14px;
}
.instr-card h4 { color: #FF9060; margin: 0 0 8px; font-size: 0.97rem; }
.instr-card ul { margin: 0; padding-left: 18px; color: #FFFFFF; font-size: 0.88rem; }
.instr-card li { margin-bottom: 4px; }

/* login */
[data-testid="stSidebar"].login-hide { display: none !important; }
[data-testid="stForm"] {
    background: rgba(10,5,2,0.75) !important;
    border: 1px solid rgba(200,80,30,0.25) !important;
    border-radius: 20px !important; padding: 10px 20px 20px !important;
}
[data-testid="stForm"] input {
    background: rgba(255,255,255,0.08) !important; color: #EDE0D4 !important;
    border: 1px solid rgba(200,80,30,0.3) !important; border-radius: 8px !important;
}
[data-testid="stForm"] label { color: #FFFFFF !important; font-weight: 600 !important; }
[data-testid="stFormSubmitButton"] > button {
    background: linear-gradient(90deg, #C8501E, #7A2E08) !important;
    color: white !important; border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 15px !important;
    padding: 10px !important; width: 100% !important;
}

#MainMenu { visibility: hidden; }
footer    { visibility: hidden; }
[data-testid="stSidebarNav"] { display: none !important; }
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# LOGO
# ─────────────────────────────────────────────────────────────────────────────
def _logo_b64() -> str:
    p = ASSETS_DIR / "ALBARID.png"
    if p.exists():
        return base64.b64encode(p.read_bytes()).decode()
    return ""

def _logo_img(size: int = 70) -> str:
    b = _logo_b64()
    if b:
        return (f'<img src="data:image/png;base64,{b}" '
                f'style="width:{size}px;height:{size}px;object-fit:contain;'
                f'border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,0.5);" />')
    return '<div style="font-size:42px;">🏦</div>'

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def _show_login() -> None:
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }
    .login-wrapper { max-width: 420px; margin: 0 auto; padding: 10px; }
    </style>
    <div class="login-wrapper">
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div style="text-align:center; padding:34px 0 20px 0;">
        {_logo_img(90)}
        <div style="font-size:26px;font-weight:900;color:#FFFFFF;
                    letter-spacing:0.5px;margin:14px 0 4px 0;">Spread Manager</div>
        <div style="font-size:16px;font-weight:700;color:#F5C518;margin-bottom:4px;">
            Al Barid Bank</div>
        <div style="font-size:12px;color:#FFFFFF;margin-bottom:20px;">
            Calculateur de Spread</div>
    </div>
    """, unsafe_allow_html=True)

    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("Username", placeholder="spreadABB",  key="lu")
        password = st.text_input("Code", type="password", placeholder="••••••••", key="lp")
        submit   = st.form_submit_button("Se connecter", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if submit:
        if CREDENTIALS.get(username) == password:
            st.session_state["authenticated"] = True
            st.session_state["username"] = username
            st.rerun()
        else:
            st.error("Identifiants incorrects. Réessayez.")

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
def _render_sidebar() -> str:
    with st.sidebar:
        st.markdown(f"""
        <div style="text-align:center;padding:14px 0 6px 0;">{_logo_img(72)}</div>
        <div style="text-align:center;padding:6px 0 16px 0;">
            <div style="font-size:15px;font-weight:900;color:#F5C518;letter-spacing:2px;">
                Al Barid Bank</div>
            <div style="font-size:11px;color:#FFFFFF;margin-top:2px;">Spread Manager</div>
            <hr style="border-color:rgba(200,80,30,0.3);margin:12px 0 8px 0;">
        </div>
        """, unsafe_allow_html=True)

        current = st.session_state.get("_nav", "Accueil")
        for item in ["Accueil", "Calculateur de Spread"]:
            css = "nav-btn-active" if item == current else "nav-btn"
            st.markdown(f'<div class="{css}">', unsafe_allow_html=True)
            if st.button(item, key=f"nav_{item}", use_container_width=True):
                st.session_state["_nav"] = item
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:rgba(200,80,30,0.2);margin:14px 0;'>",
                    unsafe_allow_html=True)
        st.markdown(
            f"<span style='font-size:11px;color:#FFFFFF;'>"
            f"Connecté : {st.session_state.get('username','')}</span>",
            unsafe_allow_html=True)
        if st.button("🔓 Déconnexion", use_container_width=True):
            st.session_state.clear()
            st.rerun()

    return st.session_state.get("_nav", "Accueil")

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _kpi(title: str, value: str, sub: str = "") -> str:
    return (f"<div class='kpi-card'><div class='kpi-title'>{title}</div>"
            f"<div class='kpi-value'>{value}</div>"
            + (f"<div class='kpi-sub'>{sub}</div>" if sub else "")
            + "</div>")

def _sec(title: str) -> None:
    st.markdown(f"<div class='section-hdr'>{title}</div>", unsafe_allow_html=True)

def _to_decimal(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().replace("%","").replace(",",".").replace(" ","")
    try:
        v = float(val)
    except (ValueError, TypeError):
        return None
    return v / 100.0 if abs(v) > 1 else v

_RATE_KW = ["COUPONRT","TAUX","RATE","RENDEMENT","INTRATE","REND","EMISSION","INTERESTRATE"]

def _detect_rate_cols(cols: list[str]) -> list[str]:
    return [c for c in cols if any(k in c.upper() for k in _RATE_KW)]

def _detect_type(row: pd.Series) -> str:
    ctgry = str(row.get("INSTRCTGRY", "")).strip().upper()
    if ctgry == "OBL_ORDN":
        return "OBLIG_ORDN"
    name = (str(row.get("ENGLONGNAME", "")) + " " + str(row.get("ENGPREFERREDNAME", ""))).upper().strip()
    first_word = name.split()[0] if name.split() else ""
    if ctgry == "TCN":
        if first_word == "BT":
            return "BT"
        if "BSF" in name:
            return "BSF"
        if "CD" in name:
            return "CD"
    return "Autre"

def _detect_sector(issuer: str) -> str:
    return SECTEUR_MAP.get(str(issuer).strip(), "AUTRES")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE ACCUEIL
# ─────────────────────────────────────────────────────────────────────────────
def _page_home() -> None:
    st.markdown(f"""
    <div style="background:linear-gradient(90deg,#1E0A03 0%,#130602 100%);
                border:1px solid rgba(200,80,30,0.4);border-radius:14px;
                padding:22px 28px;margin-bottom:22px;
                display:flex;align-items:center;gap:20px;">
        {_logo_img(60)}
        <div>
            <div style="font-size:1.55rem;font-weight:900;color:#FFFFFF;">Spread Manager</div>
            <div style="font-size:1rem;font-weight:700;color:#F5C518;">Al Barid Bank</div>
            <div style="font-size:0.85rem;color:#FFFFFF;margin-top:2px;">
                Calcul du spread BSF, CD, BT &amp; Obligations vs courbe BDT Bank Al-Maghrib</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _sec("📋 Guide d'utilisation")
    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("""
        <div class="instr-card">
            <h4>📁 Étape 1 — Charger le fichier Maroclear</h4>
            <ul>
                <li>Exportez votre fichier depuis <b>Maroclear</b> en format <b>.xlsx</b></li>
                <li>Glissez-déposez le fichier dans la zone d'upload</li>
                <li>L'application détecte automatiquement toutes les <b>feuilles disponibles</b></li>
                <li>Sélectionnez la feuille contenant les instruments</li>
            </ul>
        </div>
        <div class="instr-card">
            <h4>🔍 Étape 2 — Filtrer les instruments</h4>
            <ul>
                <li>Choisissez parmi <b>CD</b>, <b>BSF</b>, <b>BT</b> et <b>OBLIG_ORDN</b></li>
                <li>Les obligations bancaires sont automatiquement <b>exclues</b> du filtre OBLIG_ORDN</li>
                <li>Le <b>secteur</b> est détecté automatiquement pour les obligations</li>
                <li>Ajustez la <b>maturité résiduelle min/max en années</b></li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class="instr-card">
            <h4>📈 Étape 3 — Calculer les spreads</h4>
            <ul>
                <li>Récupération automatique des <b>courbes BDT BAM</b>
                    (bkam.ma) pour chaque date d'émission</li>
                <li>Les courbes téléchargées sont <b>mises en cache</b> localement</li>
                <li>Interpolation du taux BDT à la maturité exacte (logique VBA officielle BAM)</li>
                <li><b>Spread (bps) = Taux instrument − Taux BDT</b></li>
            </ul>
        </div>
        <div class="instr-card">
            <h4>📥 Étape 4 — Exporter les résultats</h4>
            <ul>
                <li>CD / BSF / BT : export groupé par type et émetteur</li>
                <li>OBLIG_ORDN : export avec une feuille globale +
                    <b>une feuille par secteur</b> (trié par spread décroissant)</li>
                <li>KPIs : spread moyen, médian, min, max — global et par type</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    _sec("🏦 Instruments couverts")
    c1, c2, c3, c4 = st.columns(4, gap="large")
    with c1:
        st.markdown("""
        <div class="instr-card">
            <h4>CD</h4>
            <ul>
                <li>Certificats de Dépôt</li>
                <li>INSTRCTGRY = TCN</li>
                <li>Maturité : 10j à 5 ans</li>
            </ul>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown("""
        <div class="instr-card">
            <h4>BSF</h4>
            <ul>
                <li>Bons de Sociétés Financières</li>
                <li>INSTRCTGRY = TCN</li>
                <li>Maturité : 2 à 5 ans</li>
            </ul>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown("""
        <div class="instr-card">
            <h4>BT</h4>
            <ul>
                <li>Billets de Trésorerie</li>
                <li>INSTRCTGRY = BDT</li>
                <li>Toutes maturités</li>
            </ul>
        </div>""", unsafe_allow_html=True)
    with c4:
        st.markdown("""
        <div class="instr-card">
            <h4>OBLIG_ORDN</h4>
            <ul>
                <li>Obligations Ordinaires</li>
                <li>INSTRCTGRY = OBL_ORDN</li>
                <li>Hors secteur bancaire</li>
            </ul>
        </div>""", unsafe_allow_html=True)

    st.info("**Connexion internet requise** pour récupérer les courbes BDT. "
            "Les courbes déjà téléchargées sont réutilisées automatiquement (cache local).")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CALCULATEUR
# ─────────────────────────────────────────────────────────────────────────────
def _page_spread() -> None:
    st.markdown(f"""
    <div style="background:linear-gradient(90deg,#1E0A03 0%,#130602 100%);
                border:1px solid rgba(200,80,30,0.4);border-radius:14px;
                padding:18px 26px;margin-bottom:20px;
                display:flex;align-items:center;gap:18px;">
        {_logo_img(52)}
        <div>
            <div style="font-size:1.35rem;font-weight:900;color:#FFFFFF;">
                Calculateur de Spread</div>
            <div style="font-size:0.85rem;color:#FFFFFF;">
                BSF, CD, BT &amp; OBLIG_ORDN — Courbe BDT Bank Al-Maghrib</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── ÉTAPE 1 : Upload ──────────────────────────────────────────────────────
    _sec("📁 Étape 1 — Charger le fichier")
    uploaded = st.file_uploader(
        "Fichier Maroclear (.xlsx)",
        type=["xlsx"],
        help="Fichier exporté depuis Maroclear contenant les instruments",
    )
    if uploaded is None:
        st.info("Chargez un fichier Excel pour commencer.")
        return

    # ── ÉTAPE 2 : Détection auto des feuilles ─────────────────────────────────
    _sec("📄 Étape 2 — Sélectionner la feuille")
    try:
        xl = pd.ExcelFile(uploaded, engine="openpyxl")
        sheet_names = xl.sheet_names
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        return

    sheet = st.selectbox(
        "Feuille Excel",
        options=sheet_names,
        help="Toutes les feuilles disponibles dans votre fichier sont listées ici.",
    )

    try:
        uploaded.seek(0)
        df = pd.read_excel(uploaded, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        st.error(f"Erreur de lecture de la feuille '{sheet}' : {e}")
        return

    st.success(f"✅ Feuille **{sheet}** chargée — {len(df)} lignes, {len(df.columns)} colonnes.")

    # ── Vérification colonnes requises ─────────────────────────────────────────
    REQUIRED = ["ISSUEDT", "MATURITYDT_L", "INSTRCTGRY", "ENGLONGNAME", "ENGPREFERREDNAME"]
    missing  = [c for c in REQUIRED if c not in df.columns]
    if missing:
        st.error(f"Colonnes manquantes : **{missing}**\n\n"
                 f"Colonnes disponibles : {list(df.columns)}")
        return

    # ── ÉTAPE 3 : Paramètres de filtrage ──────────────────────────────────────
    _sec("⚙️ Étape 3 — Paramètres de filtrage")

    # Détecter les dates min/max de ISSUEDT dans le fichier
    _issuedt_series = pd.to_datetime(df["ISSUEDT"], errors="coerce").dropna()
    _date_min_file  = _issuedt_series.min().date() if not _issuedt_series.empty else date(2020, 1, 1)
    _date_max_file  = _issuedt_series.max().date() if not _issuedt_series.empty else date.today()

    # Ligne 1 — Types d'instruments (4 checkboxes)
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        inc_cd    = st.checkbox("CD — Certificats de Dépôt",         value=True)
    with fc2:
        inc_bsf   = st.checkbox("BSF — Bons de Soc. Financières",    value=True)
    with fc3:
        inc_bt    = st.checkbox("BT — Billets de Trésorerie",         value=False)
    with fc4:
        inc_oblig = st.checkbox("OBLIG_ORDN — Obligations (hors banques)", value=False)

    # Ligne 2 — Maturité
    mc1, mc2 = st.columns(2)
    with mc1:
        res_min_y = st.number_input("Maturité min (ans)", value=0.0, min_value=0.0,
                                    max_value=30.0, step=0.5, format="%.1f")
    with mc2:
        res_max_y = st.number_input("Maturité max (ans)", value=5.0, min_value=0.0,
                                    max_value=30.0, step=0.5, format="%.1f")

    # Ligne 3 — Date d'émission (ISSUEDT)
    st.markdown(
        "<span style='font-size:0.85rem;color:#FFFFFF;font-weight:600;'>"
        "📅 Filtrer par date d'émission (ISSUEDT)</span>",
        unsafe_allow_html=True,
    )
    _bound_min = min(_date_min_file, date(1990, 1, 1))
    _bound_max = max(_date_max_file, date.today())

    dc1, dc2 = st.columns(2)
    with dc1:
        date_iss_min = st.date_input(
            "Date d'émission — Début",
            value=_date_min_file,
            min_value=_bound_min,
            max_value=_bound_max,
            key="iss_min",
        )
    with dc2:
        date_iss_max = st.date_input(
            "Date d'émission — Fin",
            value=_date_max_file,
            min_value=_bound_min,
            max_value=_bound_max,
            key="iss_max",
        )

    # Conversion maturité en jours
    res_min = int(res_min_y * 365)
    res_max = int(res_max_y * 365) if res_max_y > 0 else 36500

    if not any([inc_cd, inc_bsf, inc_bt, inc_oblig]):
        st.warning("Sélectionnez au moins un type d'instrument.")
        return

    # ── Filtrage ───────────────────────────────────────────────────────────────
    dff = df.copy()
    dff["_idt"] = pd.to_datetime(dff["ISSUEDT"],      errors="coerce").dt.date
    dff["_mdt"] = pd.to_datetime(dff["MATURITYDT_L"], errors="coerce").dt.date

    instrctgry  = dff["INSTRCTGRY"].fillna("").astype(str).str.strip().str.upper()
    name_mix    = (dff["ENGLONGNAME"].fillna("").astype(str)
                   + " " + dff["ENGPREFERREDNAME"].fillna("").astype(str)).str.upper()

    # Filtre date d'émission + maturité résiduelle (communs à tous les types)
    mask_issuedt = dff["_idt"].between(date_iss_min, date_iss_max)
    residual     = (pd.to_datetime(dff["MATURITYDT_L"], errors="coerce")
                    - pd.to_datetime(dff["ISSUEDT"],    errors="coerce")).dt.days
    mask_resid   = residual.between(res_min, res_max)

    # Construction du masque combiné
    mask_combined = pd.Series(False, index=dff.index)

    if inc_cd or inc_bsf:
        mask_tcn = instrctgry.eq("TCN")
        if inc_cd:
            mask_combined |= mask_tcn & name_mix.str.contains("CD", regex=False)
        if inc_bsf:
            mask_combined |= mask_tcn & name_mix.str.contains("BSF", regex=False)

    if inc_bt:
        name_first = dff["ENGLONGNAME"].fillna("").astype(str).str.strip().str.upper().str.split().str[0]
        mask_combined |= instrctgry.eq("TCN") & name_first.eq("BT")

    if inc_oblig:
        mask_oblig = instrctgry.eq("OBL_ORDN")
        # Exclure les émetteurs bancaires si la colonne existe
        if "PREFERREDNAMEISSUER" in dff.columns:
            mask_oblig &= ~dff["PREFERREDNAMEISSUER"].fillna("").astype(str).isin(BANK_ISSUERS)
        mask_combined |= mask_oblig

    mask_combined &= mask_issuedt & mask_resid
    selected_idx = df.index[mask_combined].tolist()

    # ── KPIs avant calcul ─────────────────────────────────────────────────────
    type_labels = []
    if inc_cd:    type_labels.append("CD")
    if inc_bsf:   type_labels.append("BSF")
    if inc_bt:    type_labels.append("BT")
    if inc_oblig: type_labels.append("OBLIG")

    type_counts: dict[str, int] = {}
    if inc_cd:
        type_counts["CD"] = int((mask_combined & instrctgry.eq("TCN") & name_mix.str.contains("CD", regex=False)).sum())
    if inc_bsf:
        type_counts["BSF"] = int((mask_combined & instrctgry.eq("TCN") & name_mix.str.contains("BSF", regex=False)).sum())
    if inc_bt:
        type_counts["BT"] = int((mask_combined & instrctgry.eq("BDT")).sum())
    if inc_oblig:
        _m_oblig = mask_combined & instrctgry.eq("OBL_ORDN")
        type_counts["OBLIG"] = int(_m_oblig.sum())

    k1, k2, k3, k4 = st.columns(4)
    k1.markdown(_kpi("Lignes totales",      str(len(df))),            unsafe_allow_html=True)
    k2.markdown(_kpi("Instruments retenus", str(len(selected_idx))),  unsafe_allow_html=True)
    tstr = " / ".join(f"{k}:{v}" for k, v in type_counts.items())
    k3.markdown(_kpi("Répartition", tstr),                            unsafe_allow_html=True)
    k4.markdown(_kpi("Période émission",
                     f"{date_iss_min.strftime('%d/%m/%Y')}",
                     f"→ {date_iss_max.strftime('%d/%m/%Y')}"),       unsafe_allow_html=True)

    if not selected_idx:
        st.warning("Aucun instrument trouvé avec ces filtres.")
        return

    # ── Colonne taux instrument (optionnelle) ──────────────────────────────────
    rate_candidates = _detect_rate_cols(list(df.columns))
    rate_col: Optional[str] = None
    if rate_candidates:
        choice   = st.selectbox(
            "Colonne taux instrument (pour calcul spread)",
            ["— aucune —"] + rate_candidates,
            index=1,
        )
        rate_col = None if choice == "— aucune —" else choice
    else:
        st.info("Aucune colonne taux détectée — seul le Taux BDT sera calculé.")

    # ── Bouton calcul ──────────────────────────────────────────────────────────
    if not st.button("⚡ Calculer les spreads", use_container_width=True):
        return

    # ── Préparer le df filtré ──────────────────────────────────────────────────
    df_work = df.loc[selected_idx].copy()

    # Déduplication sur INSTRID (garder la première occurrence)
    _dup_col = next(
        (c for c in df_work.columns
         if c.upper() in {"INSTRID", "ISIN", "ISINCODE", "CODEISIN", "INSTRISINOCODE"}),
        None,
    )
    if _dup_col:
        _n_before = len(df_work)
        df_work = df_work.drop_duplicates(subset=[_dup_col], keep="first")
        _n_after = len(df_work)
        if _n_before > _n_after:
            st.caption(f"ℹ️ {_n_before - _n_after} doublon(s) supprimé(s) sur la colonne **{_dup_col}**.")

    df_work["ISSUEDT"]      = pd.to_datetime(df_work["ISSUEDT"],      errors="coerce")
    df_work["MATURITYDT_L"] = pd.to_datetime(df_work["MATURITYDT_L"], errors="coerce")

    df_work["Maturité (ans)"] = (
        (df_work["MATURITYDT_L"] - df_work["ISSUEDT"]).dt.days / 365.25
    ).round(2)

    unique_dates = sorted({d.date() for d in df_work["ISSUEDT"].dropna()})

    # ── Récupération courbes BAM ───────────────────────────────────────────────
    _sec("🌐 Récupération des courbes BDT BAM")
    pbar    = st.progress(0.0)
    stxt    = st.empty()
    fetcher = BamCurveFetcher(cache_dir=str(CACHE_DIR))

    total_dates = len(unique_dates)

    def _progress(done: int, total: int, n_cache: int, n_net: int, eta: float) -> None:
        pbar.progress(done / max(total, 1))
        eta_str = f" — ETA : {eta:.0f}s" if eta > 1 else ""
        stxt.markdown(
            f"<span style='color:#FFFFFF;font-size:0.85rem;'>"
            f"Courbes BAM : <b>{done}/{total}</b>"
            f" &nbsp;(Supabase/cache : {n_cache} | réseau BAM : {n_net}){eta_str}</span>",
            unsafe_allow_html=True,
        )

    curves = fetcher.get_curves_parallel(
        unique_dates,
        max_workers=10,
        progress_callback=_progress,
    )

    # Pour les dates sans courbe, chercher la date ouvrable la plus proche (±5 jours)
    missing_dates = [d for d in unique_dates if not curves.get(d)]
    if missing_dates:
        fallback_candidates: set[date] = set()
        for d in missing_dates:
            for delta in range(1, 6):
                fallback_candidates.add(d - timedelta(days=delta))
                fallback_candidates.add(d + timedelta(days=delta))
        fallback_candidates -= set(unique_dates)
        fallback_curves = fetcher.get_curves_parallel(list(fallback_candidates), max_workers=10)
        for d in missing_dates:
            for delta in range(1, 6):
                for candidate in (d - timedelta(days=delta), d + timedelta(days=delta)):
                    if fallback_curves.get(candidate):
                        curves[d] = fallback_curves[candidate]
                        break
                if curves.get(d):
                    break

    pbar.empty(); stxt.empty()
    ok     = sum(1 for v in curves.values() if v is not None)
    no_data = total_dates - ok
    st.success(f"✅ {ok} courbes BDT récupérées (Supabase + cache + BAM)."
               + (f"  ❌ {no_data} date(s) sans données BAM (jour non ouvré)." if no_data else ""))

    # ── Calcul taux BDT + spread ───────────────────────────────────────────────
    bdt_rates:   list[Optional[float]] = []
    instr_rates: list[Optional[float]] = []
    spreads_bps: list[Optional[float]] = []

    for _, row in df_work.iterrows():
        idt = row["ISSUEDT"]
        mdt = row["MATURITYDT_L"]

        if pd.isna(idt) or pd.isna(mdt):
            bdt_rates.append(None); instr_rates.append(None); spreads_bps.append(None)
            continue

        mat_days = int((mdt - idt).days)
        if mat_days <= 0:
            bdt_rates.append(None); instr_rates.append(None); spreads_bps.append(None)
            continue

        curve = curves.get(idt.date())
        if not curve:
            bdt_rates.append(None); instr_rates.append(None); spreads_bps.append(None)
            continue

        mt, tx = curve
        try:
            bdt = calcul_taux(mat_days, mt, tx, idt.date())
            bdt_rates.append(bdt)
        except Exception:
            bdt_rates.append(None); instr_rates.append(None); spreads_bps.append(None)
            continue

        ir     = _to_decimal(row.get(rate_col)) if rate_col else None
        spread = (ir - bdt) * 10_000 if (ir is not None and bdt is not None) else None
        instr_rates.append(ir)
        spreads_bps.append(spread)

    # ── Attacher les colonnes calculées ───────────────────────────────────────
    df_work["Taux BDT"] = bdt_rates
    if rate_col:
        df_work["Taux instrument"] = instr_rates
        df_work["Spread (bps)"]   = spreads_bps

    # Détection du type via INSTRCTGRY
    df_work["Type"] = df_work.apply(_detect_type, axis=1)

    # Détection du secteur pour les OBLIG_ORDN
    if inc_oblig and "PREFERREDNAMEISSUER" in df_work.columns:
        df_work["SECTEUR"] = df_work.apply(
            lambda r: _detect_sector(r.get("PREFERREDNAMEISSUER", ""))
            if r["Type"] == "OBLIG_ORDN" else "",
            axis=1,
        )
    elif inc_oblig:
        df_work["SECTEUR"] = df_work["Type"].apply(
            lambda t: "AUTRES" if t == "OBLIG_ORDN" else ""
        )

    # ── KPIs spread ───────────────────────────────────────────────────────────
    if rate_col:
        valid = [s for s in spreads_bps if s is not None]
        if valid:
            _sec("📈 KPIs — Spread")
            s1, s2, s3, s4 = st.columns(4)
            s1.markdown(_kpi("Spread moyen",  f"{sum(valid)/len(valid):.1f}", "bps"), unsafe_allow_html=True)
            s2.markdown(_kpi("Spread médian", f"{median(valid):.1f}",         "bps"), unsafe_allow_html=True)
            s3.markdown(_kpi("Spread max",    f"{max(valid):.1f}",            "bps"), unsafe_allow_html=True)
            s4.markdown(_kpi("Spread min",    f"{min(valid):.1f}",            "bps"), unsafe_allow_html=True)

            for t in df_work["Type"].unique():
                sub = df_work[df_work["Type"] == t]["Spread (bps)"].dropna().tolist()
                if sub:
                    st.markdown(
                        f"<span style='color:#FFFFFF;font-size:0.85rem;'>"
                        f"<b style='color:#FF9060;'>{t}</b> — "
                        f"Spread moyen : <b>{sum(sub)/len(sub):.1f} bps</b> | "
                        f"Médian : <b>{median(sub):.1f} bps</b> | "
                        f"n = {len(sub)}</span>",
                        unsafe_allow_html=True,
                    )

    # ── Tableau ───────────────────────────────────────────────────────────────
    _sec("📋 Tableau des résultats")
    base_cols  = ["ISSUEDT", "MATURITYDT_L", "Maturité (ans)", "ENGLONGNAME", "Type", "INSTRCTGRY"]
    if inc_oblig and "SECTEUR" in df_work.columns:
        base_cols.append("SECTEUR")
    if inc_oblig and "PREFERREDNAMEISSUER" in df_work.columns:
        base_cols.insert(4, "PREFERREDNAMEISSUER")
    extra_cols = (["Taux instrument", "Taux BDT", "Spread (bps)"] if rate_col else ["Taux BDT"])
    disp = [c for c in base_cols + extra_cols if c in df_work.columns]
    # Dédupliquer en gardant l'ordre
    seen: set[str] = set()
    disp = [c for c in disp if not (c in seen or seen.add(c))]  # type: ignore[func-returns-value]

    fmt: dict = {}
    if "Maturité (ans)"   in disp: fmt["Maturité (ans)"]   = "{:.2f}"
    if "Taux BDT"         in disp: fmt["Taux BDT"]         = "{:.4%}"
    if "Taux instrument"  in disp: fmt["Taux instrument"]  = "{:.4%}"
    if "Spread (bps)"     in disp: fmt["Spread (bps)"]     = "{:.1f}"

    st.dataframe(
        df_work[disp].reset_index(drop=True).style.format(fmt),
        use_container_width=True,
        height=420,
    )

    # ═════════════════════════════════════════════════════════════════════════
    # EXPORT EXCEL
    # ═════════════════════════════════════════════════════════════════════════
    import re as _re
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    _YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _BLUE   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    _NAVY   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    _BOLD   = Font(bold=True)
    _WHITE_BOLD = Font(bold=True, color="FFFFFF")
    _THIN   = Side(border_style="thin", color="000000")
    _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _CENTER = Alignment(horizontal="center", vertical="center")
    _HIGHLIGHT_HDRS = {"TAUX BDT", "Spread", "TAUX D'INTERET"}

    # ─────────────────────────────────────────────────────────────────────────
    # Export A : CD / BSF / BT  (groupé par type+banque avec récaps croisés)
    # ─────────────────────────────────────────────────────────────────────────
    df_tcn_bt = df_work[df_work["Type"].isin(["CD", "BSF", "BT", "Autre"])].copy()

    if not df_tcn_bt.empty:
        _sec("📥 Export Excel — CD / BSF / BT")

        _BANK_ALIASES = {"SGMB": "SAHAM"}

        def _bank_tag(name: str) -> str:
            parts = str(name).strip().split()
            raw = parts[1].upper() if len(parts) >= 2 else "AUTRE"
            return _BANK_ALIASES.get(raw, raw)

        def _mat_label_from_name(details) -> str:
            s = "" if details is None else str(details).strip().lower()
            if not s:
                return "inconnue"
            s = _re.sub(r"\bsem\.\b", "sem", s)
            matches = _re.findall(
                r"(\d+)\s*(semaines?|semaine|sem|s|mois|ans?|an|jours?|jrs?|jr|j)\b", s
            )
            if not matches:
                return "inconnue"
            num, unit = matches[-1]
            unit = unit.lower()
            if unit in {"s", "sem"} or unit.startswith("semaine"):
                unit_norm = "semaine" if num == "1" else "semaines"
            elif unit in {"j", "jr", "jrs"} or unit.startswith("jour"):
                unit_norm = "jour" if num == "1" else "jours"
            elif unit.startswith("mois"):
                unit_norm = "mois"
            else:
                unit_norm = "an" if num == "1" else "ans"
            return f"{num} {unit_norm}"

        def _mat_sort_key(label: str):
            m = _re.match(r"^\s*(\d+)\s*(jour|jours|semaine|semaines|mois|an|ans)\s*$",
                          str(label).lower())
            if not m:
                return (99, 10**9)
            n = int(m.group(1)); u = m.group(2)
            if u.startswith("jour"):     return (0, n)
            if u.startswith("semaine"):  return (1, n)
            if u.startswith("mois"):     return (2, n)
            return (3, n)

        df_xls = df_tcn_bt.copy()
        df_xls["_bank"] = df_xls["ENGLONGNAME"].fillna("").apply(_bank_tag)

        has_spread = "Spread (bps)" in df_xls.columns
        if has_spread:
            df_xls["Spread"] = df_xls["Spread (bps)"]
            df_xls_filt = df_xls[df_xls["Spread (bps)"].between(10, 70)].copy()
        else:
            df_xls_filt = df_xls.copy()

        if "Taux instrument" in df_xls_filt.columns:
            df_xls_filt["_taux_instr_pct"] = df_xls_filt["Taux instrument"].apply(
                lambda v: round(float(v) * 100, 4) if pd.notna(v) and v is not None else None
            )

        n_export = len(df_xls_filt)
        n_total  = len(df_xls)
        if has_spread:
            st.info(f"Export CD/BSF/BT : **{n_export}** instruments avec spread entre **10 et 70 bps** "
                    f"(sur {n_total} calculés).")

        _ISIN_EXACT  = {"ISINCODE","ISIN","CODEISIN","ISIN_CODE","INSTRISINOCODE"}
        _CODE_APPROX = {"INSTRCODE","INSTRUMENTCODE","INSTRNO","NEMOCODE","NEMO",
                        "SECURITYNO","CODE","INSTRID","INSTRUMENTID","INSTRIDENTIFIER"}
        _code_col = (
            next((c for c in df_tcn_bt.columns if c.upper() in _ISIN_EXACT),  None)
            or next((c for c in df_tcn_bt.columns if c.upper() in _CODE_APPROX), None)
            or next(
                (c for c in df_tcn_bt.columns
                 if df_tcn_bt[c].dropna().astype(str).str.match(r"^MA\d{10}$").any()),
                None,
            )
        )

        _exp_cols = (
            ([_code_col] if _code_col else [])
            + ["ENGLONGNAME", "ISSUEDT", "MATURITYDT_L", "Maturité (ans)", "Taux BDT"]
            + (["Spread"]             if "Spread"             in df_xls_filt.columns else [])
            + (["_taux_instr_pct"]    if "_taux_instr_pct"    in df_xls_filt.columns else [])
        )
        _exp_cols = [c for c in _exp_cols if c in df_xls_filt.columns]

        _col_labels: dict = {
            "ENGLONGNAME":    "DETAILS DU TITRE",
            "ISSUEDT":        "DATE D'EMISSION",
            "MATURITYDT_L":   "DATE D'ECHEANCE",
            "Maturité (ans)": "Maturite residuelle",
            "Taux BDT":       "TAUX BDT",
            "Spread":         "Spread",
            "_taux_instr_pct":"TAUX D'INTERET",
        }
        if _code_col:
            _col_labels[_code_col] = "CODE"

        def _style_ws(ws, n_data: int) -> None:
            max_col = ws.max_column
            for c in range(1, max_col + 1):
                cell = ws.cell(1, c)
                cell.fill = _YELLOW; cell.font = _BOLD
                cell.alignment = _CENTER; cell.border = _BORDER
            hl_cols = {c for c in range(1, max_col + 1)
                       if str(ws.cell(1, c).value or "") in _HIGHLIGHT_HDRS}
            for r in range(2, n_data + 2):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    cell.fill   = _YELLOW if c in hl_cols else _GRAY
                    cell.border = _BORDER
            for col_letter, w in zip("ABCDEFGH", [18, 38, 14, 14, 14, 12, 10, 14]):
                ws.column_dimensions[col_letter].width = w

        def _apply_number_formats(ws, n_data: int) -> None:
            hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
            for r in range(2, n_data + 2):
                if "DATE D'EMISSION" in hdr:
                    ws.cell(r, hdr["DATE D'EMISSION"]).number_format  = "dd/mm/yyyy"
                if "DATE D'ECHEANCE" in hdr:
                    ws.cell(r, hdr["DATE D'ECHEANCE"]).number_format  = "dd/mm/yyyy"
                if "Maturite residuelle" in hdr:
                    ws.cell(r, hdr["Maturite residuelle"]).number_format = "0.00"
                if "TAUX BDT" in hdr:
                    ws.cell(r, hdr["TAUX BDT"]).number_format          = "0.00%"
                if "Spread" in hdr:
                    ws.cell(r, hdr["Spread"]).number_format             = "0"
                if "TAUX D'INTERET" in hdr:
                    ws.cell(r, hdr["TAUX D'INTERET"]).number_format     = "0.00"

        def _add_summary(ws, df_s: pd.DataFrame, n_data: int) -> None:
            start_r = n_data + 3
            for ci, hdr in enumerate(["MATURITE", "MOYENNE SPREAD", "SPREAD MAX", "SPREAD MIN"], 1):
                c = ws.cell(start_r, ci, hdr)
                c.fill = _BLUE; c.font = _BOLD
                c.alignment = _CENTER; c.border = _BORDER

            buckets: dict[str, list[float]] = {}
            for _, row in df_s.iterrows():
                lbl = _mat_label_from_name(row.get("DETAILS DU TITRE", ""))
                if lbl == "inconnue":
                    continue
                try:
                    sv = float(row.get("Spread", None))
                    if sv >= 0:
                        buckets.setdefault(lbl, []).append(sv)
                except (TypeError, ValueError):
                    pass

            rr = start_r + 1
            for lbl in sorted(buckets, key=_mat_sort_key):
                vals = buckets[lbl]
                avg = sum(vals) / len(vals)
                for ci, v in enumerate(
                    [lbl, f"{avg:.0f} bps", f"{max(vals):.0f} bps", f"{min(vals):.0f} bps"], 1
                ):
                    cell = ws.cell(rr, ci, v)
                    cell.fill = _GRAY; cell.border = _BORDER
                    cell.alignment = _CENTER
                    if ci == 1: cell.font = _BOLD
                rr += 1

        def _write_sheet_tcn(writer, df_s: pd.DataFrame, sheet_name: str) -> None:
            sn     = sheet_name[:31]
            df_out = df_s[_exp_cols].rename(columns=_col_labels).reset_index(drop=True)
            df_out.to_excel(writer, sheet_name=sn, index=False)
            ws     = writer.sheets[sn]
            n_data = len(df_out)
            _style_ws(ws, n_data)
            _apply_number_formats(ws, n_data)
            _add_summary(ws, df_out, n_data)

        # BSF grouping constants
        _CREDIT_CONSO_TAGS = {"WAFASALAF", "SOFAC", "EQDOM", "RCI", "SALAFIN", "CETELEM", "ATTIJARI"}
        _CREDIT_BAIL_TAGS  = {"MAGHREB", "MAGHREBBAIL", "SAHAM", "SOGELEASE", "WAFABAIL"}

        def _bsf_group(bank: str) -> str:
            t = bank.upper()
            if t in _CREDIT_CONSO_TAGS: return "credit_conso"
            if t in _CREDIT_BAIL_TAGS:  return "credit_bail"
            return "autres_bsf"

        _ORANGE_F    = PatternFill(start_color="C8501E", end_color="C8501E", fill_type="solid")
        _WHITE_BOLD_F = Font(bold=True, color="FFFFFF")

        def _build_cross_recap(ws, df_src, group_col, title_prefix) -> None:
            """Build 3 stacked tables (MAX / MIN / MOYENNE) on a worksheet."""
            if df_src.empty:
                return

            mats = sorted(df_src["_mat"].unique(), key=_mat_sort_key)
            groups = sorted(df_src[group_col].unique())
            n_cols = len(groups)

            tables = [
                ("RECAPULATIF DES MAX DES SPREAD",     "max"),
                ("RECAPULATIF DES MIN DES SPREAD",     "min"),
                ("RECAPULATIF DES MOYENNE DES SPREAD", "mean"),
            ]

            current_row = 1
            for table_title, agg_func in tables:
                # Title row — merged across all columns (1 label col + n_cols data cols)
                total_cols = 1 + n_cols
                title_cell = ws.cell(current_row, 1, f"{title_prefix} — {table_title}")
                title_cell.fill = _ORANGE_F
                title_cell.font = _WHITE_BOLD_F
                title_cell.alignment = _CENTER
                title_cell.border = _BORDER
                if total_cols > 1:
                    try:
                        ws.merge_cells(
                            start_row=current_row, start_column=1,
                            end_row=current_row, end_column=total_cols
                        )
                    except Exception:
                        pass
                    for c in range(2, total_cols + 1):
                        cell = ws.cell(current_row, c)
                        cell.fill = _ORANGE_F
                        cell.font = _WHITE_BOLD_F
                        cell.border = _BORDER
                current_row += 1

                # Header row
                ws.cell(current_row, 1, "MATURITE").fill = _YELLOW
                ws.cell(current_row, 1).font = _BOLD
                ws.cell(current_row, 1).alignment = _CENTER
                ws.cell(current_row, 1).border = _BORDER
                for ci, grp in enumerate(groups, 2):
                    cell = ws.cell(current_row, ci, str(grp))
                    cell.fill = _YELLOW
                    cell.font = _BOLD
                    cell.alignment = _CENTER
                    cell.border = _BORDER
                current_row += 1

                # Data rows
                for mat in mats:
                    ws.cell(current_row, 1, mat).font = _BOLD
                    ws.cell(current_row, 1).alignment = _CENTER
                    ws.cell(current_row, 1).border = _BORDER
                    for ci, grp in enumerate(groups, 2):
                        mask = (df_src["_mat"] == mat) & (df_src[group_col] == grp)
                        vals = df_src.loc[mask, "Spread"].dropna()
                        try:
                            vals_f = [float(v) for v in vals]
                        except (TypeError, ValueError):
                            vals_f = []
                        if vals_f:
                            if agg_func == "max":
                                v = max(vals_f)
                            elif agg_func == "min":
                                v = min(vals_f)
                            else:
                                v = sum(vals_f) / len(vals_f)
                            display = f"{v:.0f} BPS"
                        else:
                            display = "-"
                        cell = ws.cell(current_row, ci, display)
                        cell.alignment = _CENTER
                        cell.border = _BORDER
                    current_row += 1

                # 2 blank rows between tables
                current_row += 2

            # Auto-width columns (skip MergedCell objects)
            from openpyxl.utils import get_column_letter
            for ci in range(1, ws.max_column + 1):
                col_letter = get_column_letter(ci)
                max_len = 12
                for row_cells in ws.iter_rows(min_col=ci, max_col=ci):
                    for cell in row_cells:
                        try:
                            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                        except Exception:
                            pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Si le filtre 10-70 bps donne 0 résultats, exporter tout sans filtre
        if df_xls_filt.empty and not df_xls.empty:
            df_xls_filt = df_xls.copy()
            st.warning("⚠️ Aucun instrument avec spread entre 10 et 70 bps — export de tous les instruments.")

        # Add _mat column
        df_xls_filt = df_xls_filt.copy()
        df_xls_filt["_mat"] = df_xls_filt["ENGLONGNAME"].fillna("").apply(_mat_label_from_name)

        # Split by Type
        df_cd    = df_xls_filt[df_xls_filt["Type"] == "CD"].copy()
        df_bsf   = df_xls_filt[df_xls_filt["Type"] == "BSF"].copy()
        df_bt    = df_xls_filt[df_xls_filt["Type"] == "BT"].copy()
        df_autre = df_xls_filt[~df_xls_filt["Type"].isin(["CD", "BSF", "BT"])].copy()

        output_tcn = io.BytesIO()
        with pd.ExcelWriter(output_tcn, engine="openpyxl") as writer:
            wb = writer.book

            # CD: RECAP_CD sheet first, then per-bank sheets
            if not df_cd.empty:
                ws_recap_cd = wb.create_sheet("RECAP_CD")
                _build_cross_recap(ws_recap_cd, df_cd, "_bank", "CD")
                for bank, df_grp in df_cd.groupby("_bank"):
                    _write_sheet_tcn(writer, df_grp, f"CD_{bank}")

            # BSF: split into conso / bail / autres
            if not df_bsf.empty:
                df_bsf["_bsf_grp"] = df_bsf["_bank"].apply(_bsf_group)
                df_conso      = df_bsf[df_bsf["_bsf_grp"] == "credit_conso"].copy()
                df_bail       = df_bsf[df_bsf["_bsf_grp"] == "credit_bail"].copy()
                df_autres_bsf = df_bsf[df_bsf["_bsf_grp"] == "autres_bsf"].copy()
                if not df_conso.empty:
                    ws_rc = wb.create_sheet("RECAP_BSF_CONSO")
                    _build_cross_recap(ws_rc, df_conso, "_bank", "CRÉDIT CONSOMMATION")
                    _write_sheet_tcn(writer, df_conso, "BSF_credit_consommation")
                if not df_bail.empty:
                    ws_rb = wb.create_sheet("RECAP_BSF_BAIL")
                    _build_cross_recap(ws_rb, df_bail, "_bank", "CRÉDIT BAIL")
                    _write_sheet_tcn(writer, df_bail, "BSF_credit_bail")
                if not df_autres_bsf.empty:
                    _write_sheet_tcn(writer, df_autres_bsf, "BSF_autres")

            # BT: RECAP_BT sheet first, then per-bank sheets
            if not df_bt.empty:
                ws_rbt = wb.create_sheet("RECAP_BT")
                _build_cross_recap(ws_rbt, df_bt, "_bank", "BT")
                for bank, df_grp in df_bt.groupby("_bank"):
                    _write_sheet_tcn(writer, df_grp, f"BT_{bank}")

            # Autres
            if not df_autre.empty:
                _write_sheet_tcn(writer, df_autre, "Autres")

            # Fallback: if no sheets created
            if not wb.sheetnames or (len(wb.sheetnames) == 1 and "Sheet" in wb.sheetnames):
                pd.DataFrame(columns=["Aucune donnée"]).to_excel(writer, sheet_name="VIDE", index=False)
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

        output_tcn.seek(0)
        st.download_button(
            label="⬇️  Télécharger CD / BSF / BT (Excel)",
            data=output_tcn.getvalue(),
            file_name=f"spread_cd_bsf_bt_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ─────────────────────────────────────────────────────────────────────────
    # Export B : OBLIG_ORDN — par secteur avec récaps croisés
    # ─────────────────────────────────────────────────────────────────────────
    df_oblig = df_work[df_work["Type"] == "OBLIG_ORDN"].copy()

    if not df_oblig.empty:
        _sec("📥 Export Excel — Obligations (OBLIG_ORDN)")

        # Colonnes export obligations
        _ISIN_EXACT_O  = {"ISINCODE","ISIN","CODEISIN","ISIN_CODE","INSTRISINOCODE"}
        _CODE_APPROX_O = {"INSTRCODE","INSTRUMENTCODE","INSTRNO","NEMOCODE","NEMO",
                          "SECURITYNO","INSTRID","INSTRUMENTID","INSTRIDENTIFIER"}
        _instrid_col = (
            next((c for c in df_oblig.columns if c.upper() in _ISIN_EXACT_O), None)
            or next((c for c in df_oblig.columns if c.upper() in _CODE_APPROX_O), None)
        )

        # Noms de colonnes pour l'export
        _oblig_src_cols: list[str] = []
        if _instrid_col:
            _oblig_src_cols.append(_instrid_col)
        _oblig_src_cols.append("ENGPREFERREDNAME")
        if "PREFERREDNAMEISSUER" in df_oblig.columns:
            _oblig_src_cols.append("PREFERREDNAMEISSUER")
        if "SECTEUR" in df_oblig.columns:
            _oblig_src_cols.append("SECTEUR")
        _oblig_src_cols += ["ISSUEDT", "MATURITYDT_L"]
        if rate_col and rate_col in df_oblig.columns:
            _oblig_src_cols.append(rate_col)
        _oblig_src_cols += ["Maturité (ans)", "Taux BDT"]
        if "Spread (bps)" in df_oblig.columns:
            _oblig_src_cols.append("Spread (bps)")
        _oblig_src_cols = [c for c in _oblig_src_cols if c in df_oblig.columns]

        _oblig_rename: dict = {
            "ENGPREFERREDNAME":   "NOM_INSTRUMENT",
            "PREFERREDNAMEISSUER":"EMETTEUR",
            "SECTEUR":            "SECTEUR",
            "ISSUEDT":            "DATE_EMISSION",
            "MATURITYDT_L":       "DATE_ECHEANCE",
            "Maturité (ans)":     "MATURITE_ANS",
            "Taux BDT":           "TAUX_BDT_INTERP",
            "Spread (bps)":       "SPREAD_BPS",
        }
        if _instrid_col:
            _oblig_rename[_instrid_col] = "INSTRID"
        if rate_col:
            _oblig_rename[rate_col] = "INTERESTRATE"

        def _style_ws_oblig(ws) -> None:
            """En-têtes bleu marine gras blanc + largeurs auto."""
            max_col = ws.max_column
            max_row = ws.max_row
            for c in range(1, max_col + 1):
                cell = ws.cell(1, c)
                cell.fill      = _NAVY
                cell.font      = _WHITE_BOLD
                cell.alignment = _CENTER
                cell.border    = _BORDER
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    ws.cell(r, c).border = _BORDER
            # Auto-ajuster largeurs
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in col_cells
                )
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        def _apply_number_formats_oblig(ws) -> None:
            hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
            max_row = ws.max_row
            for r in range(2, max_row + 1):
                if "DATE_EMISSION" in hdr:
                    ws.cell(r, hdr["DATE_EMISSION"]).number_format  = "dd/mm/yyyy"
                if "DATE_ECHEANCE" in hdr:
                    ws.cell(r, hdr["DATE_ECHEANCE"]).number_format  = "dd/mm/yyyy"
                if "MATURITE_ANS" in hdr:
                    ws.cell(r, hdr["MATURITE_ANS"]).number_format   = "0.00"
                if "TAUX_BDT_INTERP" in hdr:
                    ws.cell(r, hdr["TAUX_BDT_INTERP"]).number_format = "0.00%"
                if "SPREAD_BPS" in hdr:
                    ws.cell(r, hdr["SPREAD_BPS"]).number_format      = "0.0"
                if "INTERESTRATE" in hdr:
                    ws.cell(r, hdr["INTERESTRATE"]).number_format    = "0.00%"

        def _sheet_name_oblig(secteur: str) -> str:
            return secteur.replace("/", "-")[:31]

        def _write_oblig_sheet(writer, df_s: pd.DataFrame, sheet_name: str) -> None:
            sn = _sheet_name_oblig(sheet_name)
            df_out = df_s[_oblig_src_cols].rename(columns=_oblig_rename).reset_index(drop=True)
            # Trier par SPREAD_BPS décroissant si disponible
            if "SPREAD_BPS" in df_out.columns:
                df_out = df_out.sort_values("SPREAD_BPS", ascending=False)
            df_out.to_excel(writer, sheet_name=sn, index=False)
            ws = writer.sheets[sn]
            _style_ws_oblig(ws)
            _apply_number_formats_oblig(ws)

        def _mat_bucket_oblig(mat_ans) -> str:
            """Map a numeric maturity in years to a display bucket label."""
            try:
                v = float(mat_ans)
            except (TypeError, ValueError):
                return "inconnue"
            if v < 1:
                return "mois"
            rounded = min([1, 2, 3, 4, 5, 7, 10], key=lambda x: abs(x - v))
            if rounded == 1:
                return "1 an"
            return f"{rounded} ans"

        def _build_oblig_recap_sheet(ws, df_renamed, secteur_title) -> None:
            _ORANGE_F     = PatternFill(start_color="C8501E", end_color="C8501E", fill_type="solid")
            _WHITE_BOLD_F = Font(bold=True, color="FFFFFF")
            """Build 3 stacked recap tables (MAX/MIN/MOYENNE) for obligations."""
            if df_renamed.empty:
                return

            # Add maturity bucket column
            df_work_r = df_renamed.copy()
            if "MATURITE_ANS" in df_work_r.columns:
                df_work_r["_mat_bkt"] = df_work_r["MATURITE_ANS"].apply(_mat_bucket_oblig)
            else:
                df_work_r["_mat_bkt"] = "inconnue"

            emetteurs = sorted(df_work_r["EMETTEUR"].dropna().unique()) if "EMETTEUR" in df_work_r.columns else []
            _bucket_order = ["mois", "1 an", "2 ans", "3 ans", "4 ans", "5 ans", "7 ans", "10 ans"]
            present_buckets = df_work_r["_mat_bkt"].unique()
            mat_buckets = [b for b in _bucket_order if b in present_buckets]
            extra = [b for b in present_buckets if b not in _bucket_order]
            mat_buckets += sorted(extra)

            if not emetteurs or not mat_buckets:
                return

            tables = [
                ("RECAPULATIF DES MAX DES SPREAD",     "max"),
                ("RECAPULATIF DES MIN DES SPREAD",     "min"),
                ("RECAPULATIF DES MOYENNE DES SPREAD", "mean"),
            ]

            current_row = 1
            for table_title, agg_func in tables:
                total_cols = 1 + len(mat_buckets)
                title_cell = ws.cell(current_row, 1, f"{secteur_title} — {table_title}")
                title_cell.fill = _ORANGE_F
                title_cell.font = _WHITE_BOLD_F
                title_cell.alignment = _CENTER
                title_cell.border = _BORDER
                if total_cols > 1:
                    try:
                        ws.merge_cells(
                            start_row=current_row, start_column=1,
                            end_row=current_row, end_column=total_cols
                        )
                    except Exception:
                        pass
                    for c in range(2, total_cols + 1):
                        cell = ws.cell(current_row, c)
                        cell.fill = _ORANGE_F
                        cell.font = _WHITE_BOLD_F
                        cell.border = _BORDER
                current_row += 1

                # Header row: first col = EMETTEUR, then maturity buckets
                ws.cell(current_row, 1, "EMETTEUR").fill = _YELLOW
                ws.cell(current_row, 1).font = _BOLD
                ws.cell(current_row, 1).alignment = _CENTER
                ws.cell(current_row, 1).border = _BORDER
                for ci, bkt in enumerate(mat_buckets, 2):
                    cell = ws.cell(current_row, ci, bkt)
                    cell.fill = _YELLOW
                    cell.font = _BOLD
                    cell.alignment = _CENTER
                    cell.border = _BORDER
                current_row += 1

                # Data rows: one per emetteur
                for emetteur in emetteurs:
                    ws.cell(current_row, 1, emetteur).font = _BOLD
                    ws.cell(current_row, 1).alignment = _CENTER
                    ws.cell(current_row, 1).border = _BORDER
                    for ci, bkt in enumerate(mat_buckets, 2):
                        mask = (
                            (df_work_r["EMETTEUR"] == emetteur) &
                            (df_work_r["_mat_bkt"] == bkt)
                        )
                        vals = df_work_r.loc[mask, "SPREAD_BPS"].dropna() if "SPREAD_BPS" in df_work_r.columns else pd.Series([], dtype=float)
                        try:
                            vals_f = [float(v) for v in vals]
                        except (TypeError, ValueError):
                            vals_f = []
                        if vals_f:
                            if agg_func == "max":
                                v = max(vals_f)
                            elif agg_func == "min":
                                v = min(vals_f)
                            else:
                                v = sum(vals_f) / len(vals_f)
                            display = f"{v:.0f} BPS"
                        else:
                            display = "-"
                        cell = ws.cell(current_row, ci, display)
                        cell.alignment = _CENTER
                        cell.border = _BORDER
                    current_row += 1

                # 2 blank rows between tables
                current_row += 2

            # Auto-width columns (skip MergedCell objects)
            from openpyxl.utils import get_column_letter
            for ci in range(1, ws.max_column + 1):
                col_letter = get_column_letter(ci)
                max_len = 12
                for row_cells in ws.iter_rows(min_col=ci, max_col=ci):
                    for cell in row_cells:
                        try:
                            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                        except Exception:
                            pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        st.info(f"Export OBLIG_ORDN : **{len(df_oblig)}** obligations (hors banques).")

        output_oblig = io.BytesIO()
        with pd.ExcelWriter(output_oblig, engine="openpyxl") as writer:
            wb_o = writer.book
            _write_oblig_sheet(writer, df_oblig, "TOUTES_OBLIG")
            secteur_col = "SECTEUR" if "SECTEUR" in df_oblig.columns else None
            if secteur_col:
                for secteur, df_grp in df_oblig.groupby(secteur_col):
                    if df_grp.empty:
                        continue
                    _write_oblig_sheet(writer, df_grp, str(secteur))
                    # Recap sheet for this sector
                    sn_recap = f"RECAP_{str(secteur)[:24]}"[:31]
                    ws_rec = wb_o.create_sheet(sn_recap)
                    df_out_grp = df_grp[_oblig_src_cols].rename(columns=_oblig_rename).reset_index(drop=True)
                    _build_oblig_recap_sheet(ws_rec, df_out_grp, str(secteur))
            if "Sheet" in wb_o.sheetnames and len(wb_o.sheetnames) > 1:
                del wb_o["Sheet"]

        output_oblig.seek(0)
        st.download_button(
            label="⬇️  Télécharger Obligations OBLIG_ORDN (Excel)",
            data=output_oblig.getvalue(),
            file_name=f"spread_oblig_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="auto",
    )
    st.markdown(STYLE, unsafe_allow_html=True)

    if not st.session_state.get("authenticated"):
        _show_login()
        return

    page = _render_sidebar()

    if page == "Accueil":
        _page_home()
    else:
        _page_spread()

if __name__ == "__main__":
    main()
