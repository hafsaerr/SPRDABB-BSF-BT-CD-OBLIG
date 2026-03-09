from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional, Sequence

import pandas as pd
from openpyxl import load_workbook


LOGGER = logging.getLogger(__name__)


@dataclass
class FilterConfig:
    issue_start: date = field(default_factory=lambda: date(2020, 1, 1))
    issue_end: date = field(default_factory=lambda: date(2030, 12, 31))
    maturity_start: date = field(default_factory=lambda: date(2020, 1, 1))
    maturity_end: date = field(default_factory=lambda: date(2040, 12, 31))
    residual_min_days: int = 1
    residual_max_days: int = 1830  # ~5 ans


# ─────────────────────────────────────────────────────────────────────────────
# LOAD
# ─────────────────────────────────────────────────────────────────────────────
_REQUIRED_COLS = [
    "ISSUEDT",
    "MATURITYDT_L",
    "INSTRCTGRY",
    "ENGLONGNAME",
    "ENGPREFERREDNAME",
]


def _ensure_required_columns(df: pd.DataFrame) -> None:
    missing = [c for c in _REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Colonnes manquantes dans le fichier : {missing}. "
            f"Colonnes disponibles : {list(df.columns)}"
        )


def load_sheet(path, sheet_name: str = "OBL_ORDN") -> pd.DataFrame:
    """Load an Excel sheet from a file path or a file-like object (Streamlit UploadedFile)."""
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    _ensure_required_columns(df)
    LOGGER.info("Feuille '%s' chargée : %d lignes, %d colonnes.", sheet_name, len(df), len(df.columns))
    return df


# ─────────────────────────────────────────────────────────────────────────────
# FILTER  —  BSF & CD (and any other TCN type)
# ─────────────────────────────────────────────────────────────────────────────
def filter_rows(
    df: pd.DataFrame,
    cfg: Optional[FilterConfig] = None,
    instrument_types: Sequence[str] = ("CD", "BSF"),
) -> pd.Series:
    """
    Return a boolean mask for rows matching:
      - INSTRCTGRY == 'TCN'
      - instrument name contains at least one of *instrument_types* (case-insensitive)
      - issue/maturity dates within cfg ranges
      - residual maturity within [cfg.residual_min_days, cfg.residual_max_days]

    Parameters
    ----------
    df : pd.DataFrame
        Raw Maroclear DataFrame (must contain _REQUIRED_COLS).
    cfg : FilterConfig, optional
        Date and maturity filter configuration.
    instrument_types : sequence of str
        Instrument type keywords to search in name columns, e.g. ('CD', 'BSF').
    """
    cfg = cfg or FilterConfig()
    dff = df.copy()

    dff["_ISSUEDT"] = pd.to_datetime(dff["ISSUEDT"], errors="coerce").dt.date
    dff["_MATURITYDT"] = pd.to_datetime(dff["MATURITYDT_L"], errors="coerce").dt.date

    # Name search field (ENGLONGNAME + ENGPREFERREDNAME)
    name_mix = (
        dff["ENGLONGNAME"].fillna("").astype(str)
        + " "
        + dff["ENGPREFERREDNAME"].fillna("").astype(str)
    ).str.upper()

    # 1) TCN category
    mask_tcn = dff["INSTRCTGRY"].fillna("").astype(str).str.upper().eq("TCN")

    # 2) Instrument type match (OR across types)
    mask_type = pd.Series(False, index=dff.index)
    for itype in instrument_types:
        mask_type = mask_type | name_mix.str.contains(itype.upper(), regex=False)

    # 3) Date filters
    mask_dates = (
        dff["_ISSUEDT"].between(cfg.issue_start, cfg.issue_end)
        & dff["_MATURITYDT"].between(cfg.maturity_start, cfg.maturity_end)
    )

    # 4) Residual maturity
    residual = (
        pd.to_datetime(dff["MATURITYDT_L"], errors="coerce")
        - pd.to_datetime(dff["ISSUEDT"], errors="coerce")
    ).dt.days
    mask_residual = residual.between(cfg.residual_min_days, cfg.residual_max_days)

    mask = mask_tcn & mask_type & mask_dates & mask_residual

    LOGGER.info(
        "Filtrage : total=%d | TCN=%d | type=%d | dates=%d | residuel=%d | retenu=%d",
        len(df),
        int(mask_tcn.sum()),
        int(mask_type.sum()),
        int(mask_dates.sum()),
        int(mask_residual.sum()),
        int(mask.sum()),
    )
    return mask


# ─────────────────────────────────────────────────────────────────────────────
# WRITE BACK  (optionnel — écriture dans l'Excel original)
# ─────────────────────────────────────────────────────────────────────────────
def write_rates_to_excel(
    input_path,
    output_path,
    sheet_name: str,
    rates_by_df_index: dict[int, Optional[float]],
    taux_col_name: str = "Taux BDT",
    spread_by_df_index: Optional[dict[int, Optional[float]]] = None,
    spread_col_name: str = "Spread (bps)",
    number_format: str = "0.000%",
) -> None:
    """
    Write BDT rates (and optionally spreads) into the Excel file.

    Parameters
    ----------
    input_path : path-like
        Source Excel file.
    output_path : path-like
        Destination Excel file.
    sheet_name : str
        Sheet to update.
    rates_by_df_index : dict
        Mapping {df_row_index: bdt_rate_decimal}.
    taux_col_name : str
        Column name for BDT rate.
    spread_by_df_index : dict, optional
        Mapping {df_row_index: spread_bps}.
    spread_col_name : str
        Column name for spread (bps).
    number_format : str
        Excel number format for rate column.
    """
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille introuvable : '{sheet_name}'. Feuilles disponibles : {wb.sheetnames}")
    ws = wb[sheet_name]

    headers = {
        str(ws.cell(row=1, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
    }

    # Ensure BDT rate column
    taux_col_idx = headers.get(taux_col_name)
    if taux_col_idx is None:
        taux_col_idx = ws.max_column + 1
        ws.cell(row=1, column=taux_col_idx).value = taux_col_name
        LOGGER.info("Colonne '%s' créée en position %d.", taux_col_name, taux_col_idx)

    # Ensure spread column (if provided)
    spread_col_idx = None
    if spread_by_df_index is not None:
        spread_col_idx = headers.get(spread_col_name)
        if spread_col_idx is None:
            spread_col_idx = ws.max_column + 1
            ws.cell(row=1, column=spread_col_idx).value = spread_col_name
            LOGGER.info("Colonne '%s' créée en position %d.", spread_col_name, spread_col_idx)

    for idx, val in rates_by_df_index.items():
        excel_row = int(idx) + 2  # header is row 1, df index 0 → row 2
        cell = ws.cell(row=excel_row, column=taux_col_idx)
        cell.value = None if val is None else float(val)
        if val is not None:
            cell.number_format = number_format

    if spread_by_df_index and spread_col_idx:
        for idx, val in spread_by_df_index.items():
            excel_row = int(idx) + 2
            cell = ws.cell(row=excel_row, column=spread_col_idx)
            cell.value = None if val is None else round(float(val), 2)

    wb.save(output_path)
    LOGGER.info("Fichier sauvegardé : %s", Path(output_path).resolve())
