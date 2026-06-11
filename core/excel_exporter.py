"""
excel_exporter.py — Fonctions d'export Excel pures (sans dépendance Streamlit).
Retournent des bytes prêts pour st.download_button.
"""
from __future__ import annotations

import io
import re as _re
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─── Styles ──────────────────────────────────────────────────────────────────
_YELLOW     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_GRAY       = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BLUE       = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_NAVY       = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_BOLD       = Font(bold=True)
_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_THIN       = Side(border_style="thin", color="000000")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER     = Alignment(horizontal="center", vertical="center")
_HIGHLIGHT_HDRS = {"TAUX BDT", "Spread", "TAUX D'INTERET"}


# ─── Helpers TCN/BT ──────────────────────────────────────────────────────────
_BANK_ALIASES: dict[str, str] = {"SGMB": "SAHAM"}


def _bank_tag(name: str) -> str:
    parts = str(name).strip().split()
    raw = parts[1].upper() if len(parts) >= 2 else "AUTRE"
    return _BANK_ALIASES.get(raw, raw)


def _mat_label_from_name(details) -> str:
    s = "" if details is None else str(details).strip().lower()
    if not s:
        return "inconnue"
    s = _re.sub(r"\bsem\.\b", "sem", s)
    matches = _re.findall(
        r"(\d+)\s*(semaines?|semaine|sem|s|mois|ans?|an|jours?|jrs?|jr|j)\b", s
    )
    if not matches:
        return "inconnue"
    num, unit = matches[-1]
    unit = unit.lower()
    if unit in {"s", "sem"} or unit.startswith("semaine"):
        unit_norm = "semaine" if num == "1" else "semaines"
    elif unit in {"j", "jr", "jrs"} or unit.startswith("jour"):
        unit_norm = "jour" if num == "1" else "jours"
    elif unit.startswith("mois"):
        unit_norm = "mois"
    else:
        unit_norm = "an" if num == "1" else "ans"
    return f"{num} {unit_norm}"


def _mat_sort_key(label: str):
    m = _re.match(r"^\s*(\d+)\s*(jour|jours|semaine|semaines|mois|an|ans)\s*$",
                  str(label).lower())
    if not m:
        return (99, 10**9)
    n = int(m.group(1))
    u = m.group(2)
    if u.startswith("jour"):     return (0, n)
    if u.startswith("semaine"):  return (1, n)
    if u.startswith("mois"):     return (2, n)
    return (3, n)


def _style_ws_tcn(ws, n_data: int) -> None:
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = _YELLOW
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER
    hl_cols = {c for c in range(1, max_col + 1)
               if str(ws.cell(1, c).value or "") in _HIGHLIGHT_HDRS}
    for r in range(2, n_data + 2):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = _YELLOW if c in hl_cols else _GRAY
            cell.border = _BORDER
    for col_letter, w in zip("ABCDEFGH", [18, 38, 14, 14, 14, 12, 10, 14]):
        ws.column_dimensions[col_letter].width = w


def _apply_number_formats_tcn(ws, n_data: int) -> None:
    hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    for r in range(2, n_data + 2):
        if "DATE D'EMISSION" in hdr:
            ws.cell(r, hdr["DATE D'EMISSION"]).number_format  = "dd/mm/yyyy"
        if "DATE D'ECHEANCE" in hdr:
            ws.cell(r, hdr["DATE D'ECHEANCE"]).number_format  = "dd/mm/yyyy"
        if "Maturite residuelle" in hdr:
            ws.cell(r, hdr["Maturite residuelle"]).number_format = "0.00"
        if "TAUX BDT" in hdr:
            ws.cell(r, hdr["TAUX BDT"]).number_format          = "0.00%"
        if "Spread" in hdr:
            ws.cell(r, hdr["Spread"]).number_format             = "0"
        if "TAUX D'INTERET" in hdr:
            ws.cell(r, hdr["TAUX D'INTERET"]).number_format     = "0.00"


def _add_summary(ws, df_s: pd.DataFrame, n_data: int) -> None:
    start_r = n_data + 3
    for ci, hdr in enumerate(["MATURITE", "MOYENNE SPREAD", "SPREAD MAX", "SPREAD MIN"], 1):
        c = ws.cell(start_r, ci, hdr)
        c.fill = _BLUE; c.font = _BOLD
        c.alignment = _CENTER; c.border = _BORDER

    buckets: dict[str, list[float]] = {}
    for _, row in df_s.iterrows():
        lbl = _mat_label_from_name(row.get("DETAILS DU TITRE", ""))
        if lbl == "inconnue":
            continue
        try:
            sv = float(row.get("Spread", None))
            if sv >= 0:
                buckets.setdefault(lbl, []).append(sv)
        except (TypeError, ValueError):
            pass

    rr = start_r + 1
    for lbl in sorted(buckets, key=_mat_sort_key):
        vals = buckets[lbl]
        avg = sum(vals) / len(vals)
        for ci, v in enumerate(
            [lbl, f"{avg:.0f} bps", f"{max(vals):.0f} bps", f"{min(vals):.0f} bps"], 1
        ):
            cell = ws.cell(rr, ci, v)
            cell.fill = _GRAY; cell.border = _BORDER
            cell.alignment = _CENTER
            if ci == 1:
                cell.font = _BOLD
        rr += 1


# ─── Helpers OBLIG_ORDN ──────────────────────────────────────────────────────
def _style_ws_oblig(ws) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill      = _NAVY
        cell.font      = _WHITE_BOLD
        cell.alignment = _CENTER
        cell.border    = _BORDER
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = _BORDER
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def _apply_number_formats_oblig(ws) -> None:
    hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        if "DATE_EMISSION" in hdr:
            ws.cell(r, hdr["DATE_EMISSION"]).number_format  = "dd/mm/yyyy"
        if "DATE_ECHEANCE" in hdr:
            ws.cell(r, hdr["DATE_ECHEANCE"]).number_format  = "dd/mm/yyyy"
        if "MATURITE_ANS" in hdr:
            ws.cell(r, hdr["MATURITE_ANS"]).number_format   = "0.00"
        if "TAUX_BDT_INTERP" in hdr:
            ws.cell(r, hdr["TAUX_BDT_INTERP"]).number_format = "0.00%"
        if "SPREAD_BPS" in hdr:
            ws.cell(r, hdr["SPREAD_BPS"]).number_format      = "0.0"
        if "INTERESTRATE" in hdr:
            ws.cell(r, hdr["INTERESTRATE"]).number_format    = "0.00%"


def _sheet_name_oblig(secteur: str) -> str:
    return secteur.replace("/", "-")[:31]


# ─── API publique ─────────────────────────────────────────────────────────────

def export_tcn_bt(df_work: pd.DataFrame, rate_col: Optional[str] = None) -> bytes:
    """
    Génère le fichier Excel CD / BSF / BT.
    Retourne les bytes du fichier .xlsx.
    """
    df_tcn_bt = df_work[df_work["Type"].isin(["CD", "BSF", "BT", "Autre"])].copy()
    if df_tcn_bt.empty:
        return b""

    df_tcn_bt["_bank"] = df_tcn_bt["ENGLONGNAME"].fillna("").apply(_bank_tag)
    has_spread = "Spread (bps)" in df_tcn_bt.columns
    if has_spread:
        df_tcn_bt["Spread"] = df_tcn_bt["Spread (bps)"]
        df_xls_filt = df_tcn_bt[df_tcn_bt["Spread (bps)"].between(10, 70)].copy()
    else:
        df_xls_filt = df_tcn_bt.copy()

    if "Taux instrument" in df_xls_filt.columns:
        df_xls_filt["_taux_instr_pct"] = df_xls_filt["Taux instrument"].apply(
            lambda v: round(float(v) * 100, 4) if pd.notna(v) and v is not None else None
        )

    _ISIN_EXACT  = {"ISINCODE", "ISIN", "CODEISIN", "ISIN_CODE", "INSTRISINOCODE"}
    _CODE_APPROX = {"INSTRCODE", "INSTRUMENTCODE", "INSTRNO", "NEMOCODE", "NEMO",
                    "SECURITYNO", "CODE", "INSTRID", "INSTRUMENTID", "INSTRIDENTIFIER"}
    _code_col = (
        next((c for c in df_tcn_bt.columns if c.upper() in _ISIN_EXACT),  None)
        or next((c for c in df_tcn_bt.columns if c.upper() in _CODE_APPROX), None)
    )

    _exp_cols = (
        ([_code_col] if _code_col else [])
        + ["ENGLONGNAME", "ISSUEDT", "MATURITYDT_L", "Maturité (ans)", "Taux BDT"]
        + (["Spread"]          if "Spread"          in df_xls_filt.columns else [])
        + (["_taux_instr_pct"] if "_taux_instr_pct" in df_xls_filt.columns else [])
    )
    _exp_cols = [c for c in _exp_cols if c in df_xls_filt.columns]

    _col_labels: dict = {
        "ENGLONGNAME":    "DETAILS DU TITRE",
        "ISSUEDT":        "DATE D'EMISSION",
        "MATURITYDT_L":   "DATE D'ECHEANCE",
        "Maturité (ans)": "Maturite residuelle",
        "Taux BDT":       "TAUX BDT",
        "Spread":         "Spread",
        "_taux_instr_pct": "TAUX D'INTERET",
    }
    if _code_col:
        _col_labels[_code_col] = "CODE"

    def _write_sheet(writer, df_s: pd.DataFrame, sheet_name: str) -> None:
        sn = sheet_name[:31]
        df_out = df_s[_exp_cols].rename(columns=_col_labels).reset_index(drop=True)
        df_out.to_excel(writer, sheet_name=sn, index=False)
        ws = writer.sheets[sn]
        n_data = len(df_out)
        _style_ws_tcn(ws, n_data)
        _apply_number_formats_tcn(ws, n_data)
        _add_summary(ws, df_out, n_data)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if not df_xls_filt.empty:
            _write_sheet(writer, df_xls_filt, "TOUT")
        df_xls_filt["_type_nn"] = df_xls_filt["Type"].fillna("AUTRE")
        df_xls_filt["_bank_nn"] = df_xls_filt["_bank"].fillna("AUTRE")
        for (typ, bank), df_grp in df_xls_filt.groupby(["_type_nn", "_bank_nn"]):
            if not df_grp.empty:
                _write_sheet(writer, df_grp, f"{typ}_{bank}")
    buf.seek(0)
    return buf.getvalue()


def export_oblig(
    df_oblig: pd.DataFrame,
    rate_col: Optional[str] = None,
    instrid_col: Optional[str] = None,
) -> bytes:
    """
    Génère le fichier Excel OBLIG_ORDN avec une feuille par secteur.
    Retourne les bytes du fichier .xlsx.
    """
    if df_oblig.empty:
        return b""

    # Déduplication
    if instrid_col and instrid_col in df_oblig.columns:
        df_oblig = df_oblig.drop_duplicates(subset=[instrid_col], keep="first")

    _ISIN_EXACT_O  = {"ISINCODE", "ISIN", "CODEISIN", "ISIN_CODE", "INSTRISINOCODE"}
    _CODE_APPROX_O = {"INSTRCODE", "INSTRUMENTCODE", "INSTRNO", "NEMOCODE", "NEMO",
                      "SECURITYNO", "INSTRID", "INSTRUMENTID", "INSTRIDENTIFIER"}
    _instrid = instrid_col or (
        next((c for c in df_oblig.columns if c.upper() in _ISIN_EXACT_O), None)
        or next((c for c in df_oblig.columns if c.upper() in _CODE_APPROX_O), None)
    )

    src_cols: list[str] = []
    if _instrid:
        src_cols.append(_instrid)
    src_cols.append("ENGPREFERREDNAME")
    _emetteur_col = next((c for c in ["PREFERREDNAMEREGISTRAR", "PREFERREDNAMEISSUER"] if c in df_oblig.columns), None)
    if _emetteur_col:
        src_cols.append(_emetteur_col)
    if "SECTEUR" in df_oblig.columns:
        src_cols.append("SECTEUR")
    src_cols += ["ISSUEDT", "MATURITYDT_L"]
    if rate_col and "Taux instrument" in df_oblig.columns:
        src_cols.append("Taux instrument")
    src_cols += ["Maturité (ans)", "Taux BDT"]
    if "Spread (bps)" in df_oblig.columns:
        src_cols.append("Spread (bps)")
    src_cols = [c for c in src_cols if c in df_oblig.columns]

    rename: dict = {
        "ENGPREFERREDNAME":   "NOM_INSTRUMENT",
        "SECTEUR":             "SECTEUR",
        "ISSUEDT":             "DATE_EMISSION",
        "MATURITYDT_L":        "DATE_ECHEANCE",
        "Maturité (ans)":      "MATURITE_ANS",
        "Taux BDT":            "TAUX_BDT_INTERP",
        "Spread (bps)":        "SPREAD_BPS",
        "Taux instrument":     "INTERESTRATE",
    }
    if _emetteur_col:
        rename[_emetteur_col] = "EMETTEUR"
    if _instrid:
        rename[_instrid] = "INSTRID"

    def _write_sheet(writer, df_s: pd.DataFrame, sheet_name: str) -> None:
        sn = _sheet_name_oblig(sheet_name)
        df_out = df_s[src_cols].rename(columns=rename).reset_index(drop=True)
        if "SPREAD_BPS" in df_out.columns:
            df_out = df_out.sort_values("SPREAD_BPS", ascending=False)
        df_out.to_excel(writer, sheet_name=sn, index=False)
        ws = writer.sheets[sn]
        _style_ws_oblig(ws)
        _apply_number_formats_oblig(ws)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_sheet(writer, df_oblig, "TOUTES_OBLIG")
        secteur_col = "SECTEUR" if "SECTEUR" in df_oblig.columns else None
        if secteur_col:
            for secteur, df_grp in df_oblig.groupby(secteur_col):
                if not df_grp.empty:
                    _write_sheet(writer, df_grp, str(secteur))
    buf.seek(0)
    return buf.getvalue()
